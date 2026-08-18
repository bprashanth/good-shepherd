#!/usr/bin/env python3
"""Bind canonical page geometry to an immutable agentic-primary workbook.

The primary workbook is never rewritten by this module.  It aligns rows from
the canonical page-shaped workbook to rows in the primary workbook, preserving
the primary's real sheet/row coordinates for review overlays.  Alignment is
monotonic and confidence-gated: an uncertain row is left unmapped rather than
guessed.
"""
from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


def normalize(value) -> str:
    if value is None:
        return ""
    text = str(value).casefold().strip()
    text = re.sub(r"\s+", "", text)
    return text.replace("–", "-").replace("—", "-")


def _identifier(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value)


def semantic_value(value) -> str:
    text = normalize(value)
    if text in {"x", "✓", "✔", "☑", "✅"}:
        return "<checked>"
    try:
        number = Decimal(text)
        if number.is_finite():
            return f"<number:{number.normalize()}>"
    except InvalidOperation:
        pass
    return text


@dataclass(frozen=True)
class WorkbookRow:
    sheet: str
    row: int
    values: tuple[str, ...]

    @property
    def nonempty(self) -> dict[int, str]:
        return {index: value for index, value in enumerate(self.values) if value}


@dataclass(frozen=True)
class RowBinding:
    canonical_sheet: str
    canonical_row: int
    primary_sheet: str
    primary_row: int
    score: float
    column_offset: int = 0


def workbook_rows(path: str | Path, *, excluded_sheets=()) -> list[WorkbookRow]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = []
    excluded = {name.casefold() for name in excluded_sheets}
    for sheet in workbook:
        if sheet.title.casefold() in excluded:
            continue
        for number, cells in enumerate(sheet.iter_rows(values_only=True), 1):
            values = tuple(normalize(value) for value in cells)
            if any(values):
                result.append(WorkbookRow(sheet.title, number, values))
    return result


def _weight(column: int, value: str) -> float:
    # Literal row keys and names/species are stronger anchors than the many
    # repeated zeros/ticks in form bodies. Header and metadata labels also tend
    # to occupy these leading columns.
    if not value:
        return 0.0
    return 6.0 if column == 0 else 3.0 if column == 1 else 1.0


def _similarity_values(left_values: tuple[str, ...],
                       right_values: tuple[str, ...]) -> float:
    a = {index: value for index, value in enumerate(left_values) if value}
    b = {index: value for index, value in enumerate(right_values) if value}
    if not a or not b:
        return 0.0
    total = sum(_weight(column, value) for column, value in a.items())
    total += sum(_weight(column, value) for column, value in b.items())
    exact = 0.0
    for column in set(a) & set(b):
        if a[column] == b[column]:
            exact += 2 * _weight(column, a[column])

    # Agentic and structured readers can disagree slightly on a long species,
    # title, or note while still identifying the same physical row.
    fuzzy = 0.0
    for column in set(a) & set(b):
        x, y = a[column], b[column]
        if x == y or min(len(x), len(y)) < 6:
            continue
        ratio = SequenceMatcher(None, x, y).ratio()
        if ratio >= 0.78:
            fuzzy += 0.8 * ratio * _weight(column, x)

    positional = (exact + fuzzy) / total

    # A token can move between columns in merged headers/metadata or because an
    # agent retained printed spacer columns. This contribution is deliberately
    # bounded; leading anchors below keep repeated zero-heavy rows from moving.
    left_counts, right_counts = Counter(a.values()), Counter(b.values())
    moved = sum((left_counts & right_counts).values())
    token_f1 = 2 * moved / (len(a) + len(b))

    anchor = 0.0
    first_a, first_b = a.get(0, ""), b.get(0, "")
    if first_a and first_b and _identifier(first_a) == _identifier(first_b):
        anchor += 0.34
    elif first_a and first_b and (_identifier(first_a).isdigit()
                                 or _identifier(first_b).isdigit()):
        anchor -= 0.14
    second_a, second_b = a.get(1, ""), b.get(1, "")
    if second_a and second_b:
        ratio = SequenceMatcher(None, second_a, second_b).ratio()
        if ratio >= 0.72:
            anchor += 0.22 * ratio
    return min(1.0, 0.58 * positional + 0.28 * token_f1 + anchor)


def row_match(left: WorkbookRow, right: WorkbookRow) -> tuple[float, int]:
    score = _similarity_values(left.values, right.values)
    offset = 0
    # When the primary lays tables side-by-side, compare the canonical row to
    # each equally wide primary column window. Values inside the unrelated
    # neighbouring table must not dilute an otherwise exact match.
    left_width = max(left.nonempty, default=-1) + 1
    right_width = max(right.nonempty, default=-1) + 1
    if left_width and right_width > left_width + 2:
        for start in range(0, right_width - left_width + 1):
            window = right.values[start:start + left_width]
            candidate = _similarity_values(left.values[:left_width], window)
            if candidate > score + 1e-9:
                score, offset = candidate, start
    return score, offset


def row_similarity(left: WorkbookRow, right: WorkbookRow) -> float:
    return row_match(left, right)[0]


def align_rows(canonical_rows: list[WorkbookRow], primary_rows: list[WorkbookRow],
               *, minimum_score: float = 0.38) -> list[RowBinding]:
    """Weighted LCS alignment in document order with confidence abstention."""
    n, m = len(canonical_rows), len(primary_rows)
    scores = [[0.0] * m for _ in range(n)]
    offsets = [[0] * m for _ in range(n)]
    for i, left in enumerate(canonical_rows):
        for j, right in enumerate(primary_rows):
            score, offset = row_match(left, right)
            if score >= minimum_score:
                scores[i][j] = score
                offsets[i][j] = offset

    # Reward strong matches superlinearly. Skips are free: missing printed
    # notes or extra primary rows do not force a false correspondence.
    dp = [[0.0] * (m + 1) for _ in range(n + 1)]
    take = [[False] * m for _ in range(n)]
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            best = max(dp[i - 1][j], dp[i][j - 1])
            score = scores[i - 1][j - 1]
            diagonal = dp[i - 1][j - 1] + score * score if score else -1.0
            if diagonal > best + 1e-9:
                dp[i][j] = diagonal
                take[i - 1][j - 1] = True
            else:
                dp[i][j] = best

    result = []
    i, j = n, m
    while i and j:
        score = scores[i - 1][j - 1]
        if take[i - 1][j - 1] and abs(
                dp[i][j] - (dp[i - 1][j - 1] + score * score)) < 1e-8:
            left, right = canonical_rows[i - 1], primary_rows[j - 1]
            result.append(RowBinding(left.sheet, left.row, right.sheet, right.row,
                                     round(score, 4), offsets[i - 1][j - 1]))
            i -= 1
            j -= 1
        elif dp[i - 1][j] >= dp[i][j - 1]:
            i -= 1
        else:
            j -= 1
    result.reverse()
    return result


def align_workbooks(canonical_xlsx: str | Path, primary_xlsx: str | Path,
                    *, minimum_score: float = 0.38) -> dict:
    canonical_rows = workbook_rows(canonical_xlsx, excluded_sheets=("ecology_review",))
    primary_rows = workbook_rows(primary_xlsx, excluded_sheets=("ecology_review",))
    sheet_order = list(dict.fromkeys(row.sheet for row in canonical_rows))
    segments = _primary_segments(primary_rows, len(sheet_order))

    # Match canonical pages to primary segments by content. Primary agents may
    # reorder pages, so this assignment is deliberately not monotonic.
    candidates = {}
    for sheet in sheet_order:
        page_rows = [row for row in canonical_rows if row.sheet == sheet]
        for index, segment in enumerate(segments):
            tentative = align_rows(page_rows, segment,
                                   minimum_score=minimum_score)
            candidates[(sheet, index)] = (
                sum(item.score * item.score for item in tentative), tentative)
    assigned = {}
    if len(segments) >= len(sheet_order):
        used = set()
        ranked = sorted(candidates, key=lambda key: candidates[key][0], reverse=True)
        for sheet, index in ranked:
            if sheet not in assigned and index not in used:
                assigned[sheet] = index
                used.add(index)
        for sheet in sheet_order:
            if sheet not in assigned:
                available = [index for index in range(len(segments)) if index not in used]
                assigned[sheet] = max(
                    available, key=lambda index: candidates[(sheet, index)][0])
                used.add(assigned[sheet])
    else:
        assigned = {sheet: max(range(len(segments)),
                               key=lambda index: candidates[(sheet, index)][0])
                    for sheet in sheet_order}

    # Within an assigned page segment, align each contiguous canonical block
    # independently. Side-by-side tables can then share primary row numbers.
    bindings = []
    for sheet in sheet_order:
        sheet_rows = [row for row in canonical_rows if row.sheet == sheet]
        primary_segment = segments[assigned[sheet]]
        blocks: list[list[WorkbookRow]] = []
        for row in sheet_rows:
            if not blocks or row.row != blocks[-1][-1].row + 1:
                blocks.append([])
            blocks[-1].append(row)
        for block in blocks:
            bindings.extend(align_rows(
                block, primary_segment, minimum_score=minimum_score))
    bindings.sort(key=lambda item: (sheet_order.index(item.canonical_sheet),
                                    item.canonical_row))
    return {
        "version": "formidable-primary-row-bridge-v1",
        "policy": "monotonic confidence-gated alignment; uncertain rows abstain",
        "canonical_nonempty_rows": len(canonical_rows),
        "primary_nonempty_rows": len(primary_rows),
        "mapped_rows": len(bindings),
        "canonical_row_coverage": round(
            len(bindings) / len(canonical_rows), 4) if canonical_rows else 1.0,
        "mean_score": round(
            sum(item.score for item in bindings) / len(bindings), 4) if bindings else 0.0,
        "primary_segments": [{
            "sheet": segment[0].sheet,
            "first_row": segment[0].row,
            "last_row": segment[-1].row,
        } for segment in segments],
        "page_segment_assignment": assigned,
        "bindings": [item.__dict__ for item in bindings],
    }


def _primary_segments(rows: list[WorkbookRow], expected_pages: int) -> list[list[WorkbookRow]]:
    """Split obvious agentic page blocks without assuming preserved page order."""
    result = []
    by_sheet = {}
    for row in rows:
        by_sheet.setdefault(row.sheet, []).append(row)
    marker = re.compile(r"^(?:(?:pdf|source))?page\d+\b", re.I)
    for sheet_rows in by_sheet.values():
        starts = [row.row for row in sheet_rows
                  if row.values and marker.match(row.values[0])]
        if len(starts) < 2:
            first_values = Counter(
                row.values[0] for row in sheet_rows
                if row.values and len(_identifier(row.values[0])) >= 8)
            repeated = [value for value, count in first_values.items()
                        if 2 <= count <= expected_pages]
            if repeated:
                best = min(repeated,
                           key=lambda value: abs(first_values[value] - expected_pages))
                starts = [row.row for row in sheet_rows if row.values[0] == best]
        if len(starts) < 2:
            result.append(sheet_rows)
            continue
        starts = sorted(set(starts))
        for index, start in enumerate(starts):
            end = starts[index + 1] if index + 1 < len(starts) else 10**9
            segment = [row for row in sheet_rows if start <= row.row < end]
            if segment:
                result.append(segment)
    return result or [rows]


def _label_score(label: str, observed: str) -> float:
    left, right = normalize(label), normalize(observed)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    plain_left, plain_right = _identifier(left), _identifier(right)
    if plain_left and (plain_left in plain_right or plain_right in plain_left):
        return 0.82
    ratio = SequenceMatcher(None, plain_left, plain_right).ratio()
    return ratio if ratio >= 0.62 else 0.0


def _table_column_map(table: dict, canonical_sheet, primary_sheet,
                      header_binding: dict | None, fallback_offset: int,
                      preferred_offset: int | None = None) -> dict[int, int]:
    columns = table.get("columns") or []
    if not columns:
        return {}
    offset = (preferred_offset if preferred_offset is not None else
              (header_binding or {}).get("column_offset", fallback_offset))
    primary_row = (header_binding or {}).get("primary_row")
    if primary_row is None:
        return {index + 1: index + 1 + offset for index in range(len(columns))}

    # Header text may span a parent row and a leaf-label row. Score each source
    # label against the concatenated local header context for every XLSX column.
    candidates = []
    lower = max(1, offset if preferred_offset is not None else offset + 1)
    span = (len(columns) + 3 if preferred_offset is not None
            else max(len(columns) * 2 + 3, len(columns) + 4))
    upper = min(primary_sheet.max_column, offset + span)
    for primary_column in range(lower, upper + 1):
        observed = " ".join(str(primary_sheet.cell(row, primary_column).value or "")
                            for row in range(max(1, primary_row - 2),
                                             min(primary_sheet.max_row, primary_row + 2) + 1))
        candidates.append((primary_column, observed))

    # Monotonic assignment handles repeated labels such as three separate
    # "Fallen" columns while allowing printed spacer columns in the primary.
    n, m = len(columns), len(candidates)
    dp = [[-1e9] * (m + 1) for _ in range(n + 1)]
    choice = [[False] * (m + 1) for _ in range(n + 1)]
    for j in range(m + 1):
        dp[0][j] = 0.0
    for i in range(1, n + 1):
        label = " ".join(value for value in (
            columns[i - 1].get("parent"), columns[i - 1].get("label")) if value)
        for j in range(1, m + 1):
            dp[i][j] = dp[i][j - 1]
            score = _label_score(label, candidates[j - 1][1])
            diagonal = dp[i - 1][j - 1] + score
            if diagonal > dp[i][j]:
                dp[i][j] = diagonal
                choice[i][j] = True
    mapping = {}
    i, j = n, m
    while i and j:
        if choice[i][j] and dp[i][j] >= dp[i][j - 1]:
            mapping[i] = candidates[j - 1][0]
            i -= 1
            j -= 1
        else:
            j -= 1

    # A zero-score forced assignment is less safe than the layout-derived
    # offset. Keep high-confidence label matches; fill gaps monotonically from
    # the nearest known mapping or the row-window offset.
    result = {}
    for canonical_column in range(1, n + 1):
        primary_column = mapping.get(canonical_column)
        label = " ".join(value for value in (
            columns[canonical_column - 1].get("parent"),
            columns[canonical_column - 1].get("label")) if value)
        observed = ""
        if primary_column:
            observed = " ".join(
                str(primary_sheet.cell(row, primary_column).value or "")
                for row in range(max(1, primary_row - 2),
                                 min(primary_sheet.max_row, primary_row + 2) + 1))
        if primary_column and _label_score(label, observed) >= 0.5:
            result[canonical_column] = primary_column
        else:
            result[canonical_column] = canonical_column + offset
    # When exact labels expose spacer columns, propagate those anchors rather
    # than reverting unmatched duplicate leaves to the unspaced fallback.
    anchors = sorted(result.items())
    for canonical_column in range(1, n + 1):
        if canonical_column in mapping:
            primary_column = mapping[canonical_column]
            label = " ".join(value for value in (
                columns[canonical_column - 1].get("parent"),
                columns[canonical_column - 1].get("label")) if value)
            observed = " ".join(
                str(primary_sheet.cell(row, primary_column).value or "")
                for row in range(max(1, primary_row - 2),
                                 min(primary_sheet.max_row, primary_row + 2) + 1))
            if _label_score(label, observed) >= 0.5:
                result[canonical_column] = primary_column
    # Preserve a one-to-one ordered column lattice. An unmatched leading
    # identifier such as "Column 1" versus "No." can otherwise collapse onto
    # the following Species column.
    for canonical_column in range(n - 1, 0, -1):
        if result[canonical_column] >= result[canonical_column + 1]:
            result[canonical_column] = max(1, result[canonical_column + 1] - 1)
    for canonical_column in range(2, n + 1):
        if result[canonical_column] <= result[canonical_column - 1]:
            result[canonical_column] = result[canonical_column - 1] + 1
    return result


def _bind_item(item: dict, primary_value, primary_model: str,
               *, xlsx_sheet: str, xlsx_row: int, xlsx_column: int) -> None:
    peers = list(item.get("readings") or [])
    primary = {"model": primary_model, "value": primary_value,
               "confidence": 1.0, "illegible": False,
               "bbox": item.get("bbox")}
    item["readings"] = [primary, *peers]
    item["value"] = primary_value
    item["confidence"] = 1.0
    item["xlsx_sheet"] = xlsx_sheet
    item["xlsx_row"] = xlsx_row
    item["xlsx_column"] = xlsx_column
    peer_values = [semantic_value(reading.get("value")) for reading in peers]
    observed = semantic_value(primary_value)
    consensus = len(peer_values) >= 2 and len(set(peer_values)) == 1
    if consensus and peer_values[0] != observed:
        item["status"] = "peer_consensus_disagreement"
    elif consensus:
        item["status"] = "agreement"
    elif peers:
        item["status"] = "peer_split"
    else:
        item["status"] = "primary_only"
    item["alternatives"] = list(dict.fromkeys(
        reading.get("value") for reading in peers
        if semantic_value(reading.get("value")) != observed))


def _mark_unmapped(item: dict) -> None:
    """Keep peer evidence but never present it as delivered primary content."""
    item["value"] = None
    item["confidence"] = 0.0
    item["status"] = "unmapped_primary"
    item.pop("xlsx_sheet", None)
    item["alternatives"] = list(dict.fromkeys(
        reading.get("value") for reading in item.get("readings") or []
        if semantic_value(reading.get("value"))))


def bind_primary(document: dict, canonical_xlsx: str | Path,
                 primary_xlsx: str | Path, *, primary_model="codex:agentic-low",
                 minimum_score: float = 0.38) -> dict:
    """Mutate canonical items to present immutable primary XLSX values/coords."""
    bridge = align_workbooks(canonical_xlsx, primary_xlsx,
                             minimum_score=minimum_score)
    bindings = {(item["canonical_sheet"], item["canonical_row"]): item
                for item in bridge["bindings"]}
    canonical_book = openpyxl.load_workbook(canonical_xlsx, data_only=True)
    primary_book = openpyxl.load_workbook(primary_xlsx, data_only=True)
    mapped = total = mapped_nonblank_peer = nonblank_peer = 0

    for page in document.get("pages") or []:
        canonical_sheet_name = f"page{page['page_number']}"
        if canonical_sheet_name not in canonical_book:
            continue
        canonical_sheet = canonical_book[canonical_sheet_name]
        for item in [*(page.get("metadata_fields") or []),
                     *(page.get("free_text_regions") or [])]:
            total += 1
            binding = bindings.get((canonical_sheet_name, item.get("xlsx_row")))
            if not binding:
                _mark_unmapped(item)
                continue
            column = item.get("xlsx_column", 2) + binding["column_offset"]
            primary_sheet = primary_book[binding["primary_sheet"]]
            value = primary_sheet.cell(binding["primary_row"], column).value
            _bind_item(item, value, primary_model, xlsx_sheet=primary_sheet.title,
                       xlsx_row=binding["primary_row"], xlsx_column=column)
            mapped += 1

        for table in page.get("tables") or []:
            rows = table.get("rows") or []
            if not rows:
                continue
            first_data_row = min(cell.get("xlsx_row") for row in rows
                                 for cell in row.get("cells") or []
                                 if cell.get("xlsx_row") is not None)
            header_binding = bindings.get((canonical_sheet_name, first_data_row - 1))
            title_binding = None
            title = table.get("title") or ""
            if title:
                ranked_titles = []
                for row_number in range(max(1, first_data_row - 4), first_data_row - 1):
                    observed = " ".join(str(value or "") for value in
                                        next(canonical_sheet.iter_rows(
                                            min_row=row_number, max_row=row_number,
                                            values_only=True)))
                    binding = bindings.get((canonical_sheet_name, row_number))
                    if binding:
                        ranked_titles.append((_label_score(title, observed), binding))
                if ranked_titles and max(ranked_titles, key=lambda item: item[0])[0] >= 0.5:
                    title_binding = max(ranked_titles, key=lambda item: item[0])[1]
            data_binding = next((binding for row in rows if row.get("cells")
                                 if (binding := bindings.get((
                                     canonical_sheet_name,
                                     row["cells"][0].get("xlsx_row"))))), None)
            anchor = title_binding or header_binding or data_binding
            if not anchor:
                for row in rows:
                    for cell in row.get("cells") or []:
                        total += 1
                        nonblank_peer += cell.get("value") not in (None, "")
                        _mark_unmapped(cell)
                continue
            primary_sheet = primary_book[anchor["primary_sheet"]]
            column_map = _table_column_map(
                table, canonical_sheet, primary_sheet, header_binding,
                anchor.get("column_offset", 0),
                title_binding.get("column_offset") if title_binding else None)
            table["primary_column_map"] = column_map
            for row in rows:
                cells = row.get("cells") or []
                if not cells:
                    continue
                binding = bindings.get((canonical_sheet_name,
                                        cells[0].get("xlsx_row")))
                for cell in cells:
                    total += 1
                    is_nonblank = cell.get("value") not in (None, "")
                    nonblank_peer += is_nonblank
                    if not binding:
                        _mark_unmapped(cell)
                        continue
                    primary_column = column_map.get(cell.get("xlsx_column"))
                    if not primary_column:
                        _mark_unmapped(cell)
                        continue
                    value = primary_sheet.cell(binding["primary_row"],
                                               primary_column).value
                    _bind_item(cell, value, primary_model,
                               xlsx_sheet=primary_sheet.title,
                               xlsx_row=binding["primary_row"],
                               xlsx_column=primary_column)
                    mapped += 1
                    mapped_nonblank_peer += is_nonblank

    document["models"] = [primary_model, *(document.get("models") or [])]
    bridge.update({
        "target_items": total, "mapped_items": mapped,
        "item_coverage": round(mapped / total, 4) if total else 1.0,
        "peer_nonblank_items": nonblank_peer,
        "mapped_peer_nonblank_items": mapped_nonblank_peer,
        "peer_nonblank_coverage": round(
            mapped_nonblank_peer / nonblank_peer, 4) if nonblank_peer else 1.0,
    })
    return bridge
