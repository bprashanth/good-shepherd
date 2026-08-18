#!/usr/bin/env python3
"""Wide cross-sector form-transcription benchmark harness.

Generalisation of benchmarks/model_bench.py (see FINDINGS_treeplots.md for the
seed results): form + golden are arguments, the prompt is sector-agnostic, and
two providers are added (local OpenAI-compatible endpoints, AWS Textract).
Rendering/tiling/cost/parsing/scoring are carried over unchanged from the
proven code.

A "form dir" holds input.pdf + golden.xlsx (+ provenance.md). Rendered pages,
tiles and per-model outputs land inside it:

  forms/<sector>__<name>/
      input.pdf  golden.xlsx  provenance.md
      pages/page_N.png   tiles/page_N_hK.png
      outputs/<provider>__<model>__<mode>.{txt,xlsx,json}

Usage:
  python3 wide_bench.py render  --form forms/health__opd
  python3 wide_bench.py tiles   --form forms/health__opd
  python3 wide_bench.py run     --form forms/health__opd --provider gemini \
          --model gemini-2.5-flash --mode tiled
  python3 wide_bench.py run ... --provider local --endpoint http://localhost:8010/v1
  python3 wide_bench.py run ... --provider textract --model textract --mode oneshot
"""
import argparse, base64, csv, io, json, os, sys, time, urllib.request, urllib.error
from pathlib import Path

HERE   = Path(__file__).parent
FORMID = Path.home() / "src/github.com/bprashanth/form-idable"
GSHEP  = Path.home() / "src/github.com/bprashanth/good-shepherd/agents/formidable"
RENDER = Path(os.environ.get("FORMIDABLE_RENDER_TOOL", GSHEP / "tools/render_page.py"))
CFG    = Path.home() / ".config/formidable"

sys.path.insert(0, str(HERE))
import wide_diff                          # noqa: E402  (recall + precision/F1)
import openpyxl                           # noqa: E402


def _key(name):
    env_name = f"{name.upper()}_API_KEY"
    if os.environ.get(env_name):
        return os.environ[env_name]
    return json.loads((CFG / f"{name}.json").read_text())["api_key"]


# ── sector-agnostic transcription prompt ──────────────────────────
TRANSCRIBE_PROMPT = """You are transcribing scanned paper forms into tabular text.
The forms have a printed structure (titles, labels, table grids, checkboxes)
filled in BY HAND: handwritten numbers, words, dates, ticks and tally marks.

You are given page images of one form (a page may be split into overlapping
top/bottom halves — transcribe each region once). Transcribe EVERYTHING
visible on every page: the form title, metadata header fields (dates, IDs,
names, places, signatures noted as their written name), every table (header row
plus every data row), checkbox states, tally marks, and marginal notes.

Notation rules:
- a dot/period alone in a cell means the value 0
- a continuous line/dash struck through a cell means "no entry" (leave blank)
- tally marks (I, l, |, IIII) are a count — sum them to an integer
- a tick / X / checkmark in a box or cell means checked -> transcribe as X
- an empty checkbox or empty cell -> leave blank

Output ONLY the transcription as CSV — one table row per line, cells separated
by commas; label:value pairs as "label,value" lines. Put a line "### PAGE N"
before each page's content. Do not add prose, explanations, or markdown fences.
Transcribe values as literally as you can read them; it is better to include an
uncertain value than to omit it.
"""


# ── rendering (unchanged mechanics, form-dir aware) ───────────────
def render_pages(form_dir: Path):
    import subprocess, fitz
    pdf = form_dir / "input.pdf"
    pages_dir = form_dir / "pages"; pages_dir.mkdir(parents=True, exist_ok=True)
    n = fitz.open(str(pdf)).page_count
    out = []
    for p in range(1, n + 1):
        dst = pages_dir / f"page_{p}.png"
        subprocess.run([sys.executable, str(RENDER), str(pdf),
                        "--out", str(dst), "--page", str(p), "--zoom", "3"],
                       check=True, capture_output=True)
        out.append(dst)
    print(f"rendered {len(out)} pages -> {pages_dir}")
    return out


def render_tiles(form_dir: Path):
    """Top/bottom halves near the 1568px vision cap — the deterministic
    stand-in for codex's crop/zoom (see FINDINGS_treeplots.md)."""
    import subprocess, fitz
    pdf = form_dir / "input.pdf"
    tiles_dir = form_dir / "tiles"; tiles_dir.mkdir(parents=True, exist_ok=True)
    n = fitz.open(str(pdf)).page_count
    out = []
    for p in range(1, n + 1):
        for half, (y0, y1) in enumerate([(0.0, 0.55), (0.45, 1.0)]):
            dst = tiles_dir / f"page_{p}_h{half}.png"
            subprocess.run([sys.executable, str(RENDER), str(pdf), "--out", str(dst),
                            "--page", str(p), "--bbox", f"0,{y0},1,{y1}", "--zoom", "6"],
                           check=True, capture_output=True)
            out.append(dst)
    print(f"rendered {len(out)} tiles -> {tiles_dir}")
    return out


def _images(form_dir: Path, mode: str):
    d = form_dir / ("tiles" if mode == "tiled" else "pages")
    return sorted(d.glob("page_*.png"))


def _b64(path):
    return base64.b64encode(Path(path).read_bytes()).decode()


# ── output parsing -> xlsx (unchanged) ────────────────────────────
def text_to_xlsx(text, dst):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    rows = []
    def flush(name, rws):
        if not rws: return
        ws = wb.create_sheet(title=name[:31])
        for r in rws:
            ws.append(r)
    cur = "v2"
    for line in text.splitlines():
        if line.strip().lower().startswith("### page"):
            flush(cur, rows); rows = []; cur = line.strip().split("###")[-1].strip() or "page"
            continue
        rows.append(next(csv.reader([line])))
    flush(cur, rows)
    if not wb.sheetnames:
        wb.create_sheet("v2")
    wb.save(dst)
    return dst


# ── HTTP helper ───────────────────────────────────────────────────
def _post(url, payload, headers, timeout=600):
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers={"Content-Type": "application/json", **headers})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode())


# ── providers ─────────────────────────────────────────────────────
def gemini_oneshot(model, pages, endpoint=None):
    key = _key("gemini")
    parts = [{"text": TRANSCRIBE_PROMPT}]
    for p in pages:
        parts.append({"inline_data": {"mime_type": "image/png", "data": _b64(p)}})
    # Reasoning off. The knob differs by generation and getting it wrong is
    # expensive: 2.5 takes `thinkingBudget: 0`; 3.5/3.6 REJECT that and need
    # `thinkingLevel: "minimal"`. `includeThoughts: false` only hides thinking
    # tokens, it does not stop them being generated or billed. Earlier runs
    # stripped the config entirely when 3.6 rejected it, so those models were
    # benchmarked with reasoning ON and their cost was overstated.
    gen = {"temperature": 0}
    gen["thinkingConfig"] = ({"thinkingLevel": "minimal"}
                             if model.startswith(("gemini-3", "gemini-4"))
                             else {"thinkingBudget": 0})
    payload = {"contents": [{"role": "user", "parts": parts}], "generationConfig": gen}
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    t0 = time.time()
    try:
        resp = _post(url, payload, {})
    except urllib.error.HTTPError as e:
        if e.code == 400:                       # unknown knob for this model
            payload["generationConfig"].pop("thinkingConfig", None)
            resp = _post(url, payload, {})
        else:
            raise
    dt = time.time() - t0
    text = "".join(part.get("text", "")
                   for part in resp["candidates"][0]["content"]["parts"])
    um = resp.get("usageMetadata", {})
    return text, {"in_tok": um.get("promptTokenCount"), "out_tok": um.get("candidatesTokenCount"),
                  "cost_usd": _gemini_cost(model, um), "latency_s": round(dt, 1)}


def openrouter_oneshot(model, pages, endpoint=None):
    key = _key("openrouter")
    content = [{"type": "text", "text": TRANSCRIBE_PROMPT}]
    for p in pages:
        content.append({"type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{_b64(p)}"}})
    payload = {"model": model, "temperature": 0,
               "messages": [{"role": "user", "content": content}],
               "reasoning": {"enabled": False}, "usage": {"include": True}}
    hdrs = {"Authorization": f"Bearer {key}",
            "HTTP-Referer": "https://formidable.local", "X-Title": "formidable-eval"}
    t0 = time.time()
    try:
        resp = _post("https://openrouter.ai/api/v1/chat/completions", payload, hdrs)
    except urllib.error.HTTPError as e:
        if e.code == 400 and b"easoning" in e.read():
            payload.pop("reasoning", None)
            resp = _post("https://openrouter.ai/api/v1/chat/completions", payload, hdrs)
        else:
            raise
    dt = time.time() - t0
    text = resp["choices"][0]["message"]["content"]
    usage = resp.get("usage", {})
    cost = usage.get("cost")
    if cost is None:
        cost = _openrouter_cost(resp.get("id"), key)
    return text, {"in_tok": usage.get("prompt_tokens"), "out_tok": usage.get("completion_tokens"),
                  "cost_usd": cost, "latency_s": round(dt, 1)}


def local_oneshot(model, pages, endpoint="http://localhost:8010/v1"):
    """OpenAI-compatible local endpoint (vLLM / llama.cpp / Ollama). $0/form."""
    content = [{"type": "text", "text": TRANSCRIBE_PROMPT}]
    for p in pages:
        content.append({"type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{_b64(p)}"}})
    # A LoRA-tuned small model reads well but can fail to emit EOS and run off
    # into arithmetic progressions ("37.7,6.4 / 37.8,6.4 / ..."), which destroys
    # precision while leaving recall high. Penalise repetition and cap length.
    payload = {"model": model, "temperature": 0,
               "max_tokens": int(os.environ.get("LOCAL_MAX_TOKENS", "4096")),
               "repetition_penalty": float(os.environ.get("LOCAL_REP_PENALTY", "1.05")),
               "messages": [{"role": "user", "content": content}]}
    t0 = time.time()
    resp = _post(f"{endpoint.rstrip('/')}/chat/completions", payload, {}, timeout=1800)
    dt = time.time() - t0
    text = resp["choices"][0]["message"]["content"]
    usage = resp.get("usage", {})
    return text, {"in_tok": usage.get("prompt_tokens"), "out_tok": usage.get("completion_tokens"),
                  "cost_usd": 0.0, "latency_s": round(dt, 1)}


# Textract: $0.015/page (tables). Detection-only pipeline — no LLM.
TEXTRACT_PER_PAGE_TABLES = 0.015
def textract_oneshot(model, pages, endpoint=None):
    import boto3
    client = boto3.client("textract", region_name="ap-south-1")
    t0 = time.time()
    all_lines = []
    for i, p in enumerate(pages, 1):
        resp = client.analyze_document(
            Document={"Bytes": Path(p).read_bytes()}, FeatureTypes=["TABLES"])
        blocks = {b["Id"]: b for b in resp["Blocks"]}
        all_lines.append(f"### PAGE {i}")
        in_cells = set()
        for b in resp["Blocks"]:
            if b["BlockType"] != "TABLE":
                continue
            grid = {}
            for rel in b.get("Relationships", []):
                if rel["Type"] != "CHILD": continue
                for cid in rel["Ids"]:
                    cell = blocks[cid]
                    if cell["BlockType"] != "CELL": continue
                    words = []
                    for crel in cell.get("Relationships", []):
                        if crel["Type"] != "CHILD": continue
                        for wid in crel["Ids"]:
                            w = blocks[wid]
                            if w["BlockType"] == "WORD":
                                words.append(w["Text"]); in_cells.add(wid)
                            elif w["BlockType"] == "SELECTION_ELEMENT":
                                if w.get("SelectionStatus") == "SELECTED":
                                    words.append("X")
                    grid.setdefault(cell["RowIndex"], {})[cell["ColumnIndex"]] = " ".join(words)
            for r in sorted(grid):
                row = grid[r]
                all_lines.append(",".join(
                    '"%s"' % row.get(c, "").replace('"', "'")
                    for c in range(1, max(row) + 1)))
        # non-table text (headers, notes): LINE blocks whose words aren't in cells
        for b in resp["Blocks"]:
            if b["BlockType"] != "LINE": continue
            wids = [i2 for rel in b.get("Relationships", []) if rel["Type"] == "CHILD"
                    for i2 in rel["Ids"]]
            if wids and all(w not in in_cells for w in wids):
                all_lines.append('"%s"' % b["Text"].replace('"', "'"))
    dt = time.time() - t0
    return "\n".join(all_lines), {"in_tok": None, "out_tok": None,
                                  "cost_usd": round(TEXTRACT_PER_PAGE_TABLES * len(pages), 4),
                                  "latency_s": round(dt, 1)}


# ── cost helpers (unchanged) ──────────────────────────────────────
_GEMINI_PRICES = {
    "gemini-2.5-flash":       (0.30, 2.50),
    "gemini-2.0-flash":       (0.10, 0.40),
    "gemini-flash-latest":    (0.30, 2.50),
    # July 2026 list prices (devtk.ai / pricepertoken.com)
    "gemini-3.5-flash":       (1.50, 9.00),
    "gemini-3.6-flash":       (1.50, 7.50),
    "gemini-3.1-pro-preview": (2.00, 12.00),
    "gemini-3.5-flash-lite":  (0.30, 2.50),
}
def _gemini_cost(model, um):
    pin, pout = _GEMINI_PRICES.get(model, (0.30, 2.50))
    it = (um.get("promptTokenCount") or 0) / 1e6
    ot = (um.get("candidatesTokenCount") or 0) / 1e6
    return round(it * pin + ot * pout, 5)


def _openrouter_cost(gen_id, key):
    if not gen_id: return None
    try:
        time.sleep(1)
        req = urllib.request.Request(
            f"https://openrouter.ai/api/v1/generation?id={gen_id}",
            headers={"Authorization": f"Bearer {key}"})
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode())["data"].get("total_cost")
    except Exception:
        return None


PROVIDERS = {"gemini": gemini_oneshot, "openrouter": openrouter_oneshot,
             "local": local_oneshot, "textract": textract_oneshot}


# ── driver ────────────────────────────────────────────────────────
PROMPT_OVERRIDE = None      # set by --prompt-file; wins in every mode

PAGE_PROMPT = """You are transcribing ONE page of a scanned hand-filled paper form
(printed structure filled in by hand). You get the page as two overlapping
top/bottom half images — transcribe each region once.

Transcribe EVERYTHING: title, header fields, every table row incl. header,
checkbox states, tally marks, marginal notes.
Notation: lone dot=0; line struck through cell=blank; tally marks sum to an
integer; tick/X=X; empty box=blank.
Output ONLY CSV (cells comma-separated, label:value pairs as "label,value").
No prose, no markdown fences. Better an uncertain value than an omission.
"""


def _send_perpage(sender, model, form_dir, endpoint):
    """One request per page (both tiles of that page). Keeps multi-page forms
    inside provider image limits and matches how the goldens were converted."""
    global TRANSCRIBE_PROMPT
    tiles = sorted((form_dir / "tiles").glob("page_*_h*.png"))
    pages = sorted({int(t.stem.split("_")[1]) for t in tiles})
    # --prompt-file must win here too. It used to be clobbered by PAGE_PROMPT,
    # so every perpage prompt experiment silently ran the default prompt.
    saved, TRANSCRIBE_PROMPT = TRANSCRIBE_PROMPT, (PROMPT_OVERRIDE or PAGE_PROMPT)
    parts, cost, in_tok, out_tok, t0 = [], 0.0, 0, 0, time.time()
    try:
        for p in pages:
            pt = [t for t in tiles if t.stem.startswith(f"page_{p}_")]
            last = None
            for attempt in range(3):
                try:
                    text, meta = sender(model, pt, endpoint=endpoint)
                    last = None
                    break
                except Exception as e:  # noqa: BLE001
                    last = e
                    time.sleep(4)
            if last is not None:
                raise last
            cost += meta.get("cost_usd") or 0
            in_tok += meta.get("in_tok") or 0
            out_tok += meta.get("out_tok") or 0
            parts.append(f"### PAGE {p}\n{text}")
    finally:
        TRANSCRIBE_PROMPT = saved
    return "\n".join(parts), {"in_tok": in_tok, "out_tok": out_tok,
                              "cost_usd": round(cost, 5),
                              "latency_s": round(time.time() - t0, 1)}


def run_one(form_dir: Path, provider, model, mode="tiled", endpoint=None):
    sender = PROVIDERS.get(provider)
    if not sender:
        raise SystemExit(f"unsupported provider {provider}")
    if mode == "perpage":
        images = sorted((form_dir / "tiles").glob("page_*_h*.png"))
    else:
        images = _images(form_dir, mode)
    assert images, f"run `render`/`tiles` first for {form_dir}"
    out_dir = form_dir / "outputs"; out_dir.mkdir(parents=True, exist_ok=True)
    tag = f"{provider}__{model.replace('/', '_')}__{mode}"
    try:
        if mode == "perpage":
            text, meta = _send_perpage(sender, model, form_dir, endpoint)
        else:
            text, meta = sender(model, images, endpoint=endpoint)
    except urllib.error.HTTPError as e:
        err = {"form": form_dir.name, "model": model, "provider": provider, "mode": mode,
               "error": f"HTTP {e.code}: {e.read().decode()[:300]}"}
        print(json.dumps(err)); return err
    (out_dir / f"{tag}.txt").write_text(text)
    xlsx = text_to_xlsx(text, out_dir / f"{tag}.xlsx")
    result = wide_diff.compare(str(form_dir / "golden.xlsx"), str(xlsx))
    m = result["metrics"]
    row = {"form": form_dir.name, "model": model, "provider": provider, "mode": mode,
           "passed": result["passed"]}
    # carry EVERY metric the scorer emits; a hard-coded list silently dropped
    # the code_*/all_* buckets when they were added and made v3 look broken
    row.update({k: v for k, v in m.items()
                if k not in ("golden_cells", "candidate_cells")})
    row.update(meta)
    (out_dir / f"{tag}.json").write_text(json.dumps(row, indent=2))
    print(json.dumps(row))
    return row


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["render", "tiles", "run"])
    ap.add_argument("--form", required=True, help="form dir with input.pdf + golden.xlsx")
    ap.add_argument("--provider"); ap.add_argument("--model")
    ap.add_argument("--mode", default="tiled", choices=["oneshot", "tiled", "perpage"])
    ap.add_argument("--endpoint", default=None, help="base URL for --provider local")
    ap.add_argument("--prompt-file", default=None,
                    help="override TRANSCRIBE_PROMPT (e.g. OCR-specialist models)")
    a = ap.parse_args()
    if a.prompt_file:
        TRANSCRIBE_PROMPT = PROMPT_OVERRIDE = Path(a.prompt_file).read_text()
    form_dir = Path(a.form).resolve()
    if a.cmd == "render":
        render_pages(form_dir)
    elif a.cmd == "tiles":
        render_tiles(form_dir)
    else:
        assert a.provider and a.model, "--provider and --model required for run"
        run_one(form_dir, a.provider, a.model, a.mode, a.endpoint)
