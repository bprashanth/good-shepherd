#!/usr/bin/env python3
"""Separate ecology QA stage: evidence-backed flags, never silent correction."""
from __future__ import annotations

import argparse
import concurrent.futures
import hashlib
import json
import math
import os
import re
import statistics
import sys
import urllib.parse
import urllib.request
from dataclasses import dataclass, asdict
from pathlib import Path

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import canonical  # noqa: E402

GBIF_MATCH = "https://api.gbif.org/v1/species/match"
GBIF_OCCURRENCE = "https://api.gbif.org/v1/occurrence/search"


@dataclass
class Record:
    location: dict
    label: str
    value: object


class GBIFClient:
    def __init__(self, cache: Path):
        self.cache = cache
        cache.mkdir(parents=True, exist_ok=True)

    def get(self, url, params):
        query = urllib.parse.urlencode(params)
        full_url = f"{url}?{query}"
        key = hashlib.sha256(full_url.encode()).hexdigest()
        path = self.cache / f"{key}.json"
        if path.exists():
            return json.loads(path.read_text()), full_url
        request = urllib.request.Request(full_url, headers={"User-Agent": "Formidable-ecology-review/1"})
        with urllib.request.urlopen(request, timeout=30) as response:
            result = json.load(response)
        path.write_text(json.dumps(result, indent=2))
        return result, full_url

    def match(self, name):
        return self.get(GBIF_MATCH, {"name": name, "strict": "false", "verbose": "true"})

    def nearby_count(self, taxon_key, latitude, longitude, radius_degrees=1.0):
        south, north = max(-90, latitude - radius_degrees), min(90, latitude + radius_degrees)
        west, east = max(-180, longitude - radius_degrees), min(180, longitude + radius_degrees)
        polygon = (f"POLYGON(({west} {south},{east} {south},{east} {north},"
                   f"{west} {north},{west} {south}))")
        return self.get(GBIF_OCCURRENCE,
                        {"taxon_key": taxon_key, "geometry": polygon, "limit": 0})


def edit_distance(a, b):
    a, b = a.casefold(), b.casefold()
    previous = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        current = [i]
        for j, cb in enumerate(b, 1):
            current.append(min(current[-1] + 1, previous[j] + 1,
                               previous[j - 1] + (ca != cb)))
        previous = current
    return previous[-1]


def kind_of(label):
    text = " ".join(str(label).casefold().split())
    if any(term in text for term in ("scientific name", "species", "taxon", "spp")):
        return "species"
    if "latitude" in text or re.search(r"\blat\b", text):
        return "latitude"
    if "longitude" in text or re.search(r"\b(?:lon|long)\b", text):
        return "longitude"
    if re.search(r"\bph\b", text):
        return "ph"
    if any(term in text for term in ("temperature", "temp", "°c")):
        return "temperature"
    if any(term in text for term in ("percent", "%", "cover")):
        return "percent"
    if any(term in text for term in ("height", "diameter", "dbh", "gbh", "weight",
                                     "length", "width", "depth", "mass", "rainfall")):
        return "measurement"
    if any(term in text for term in ("count", "number", "quantity", "abundance")):
        return "count"
    return None


def number(value):
    if isinstance(value, bool) or value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    result = float(text)
    return result if math.isfinite(result) else None


def finding(record, code, severity, message, **extra):
    return {"code": code, "severity": severity, "message": message,
            "location": record.location, "label": record.label,
            "observed": record.value, **extra}


def taxon_query(value):
    """Choose one taxon-shaped line from a cell without inventing a name.

    Survey cells often retain both a common name and a scientific binomial.
    The catalogue should check the written binomial, not concatenate the two.
    Single-line values are left literal so common-name matching still works.
    """
    lines = [" ".join(line.split()) for line in str(value or "").splitlines()
             if line.strip()]
    if len(lines) > 1:
        binomials = [line for line in lines if re.fullmatch(
            r"[A-Z][a-z]+(?:\s+[a-z][a-z.-]+){1,2}", line)]
        if binomials:
            return binomials[-1]
    return " ".join(str(value or "").split())


def numeric_findings(records):
    findings = []
    hard_bounds = {"percent": (0, 100), "ph": (0, 14), "latitude": (-90, 90),
                   "longitude": (-180, 180), "temperature": (-100, 100)}
    groups = {}
    for record in records:
        kind, value = kind_of(record.label), number(record.value)
        if kind and value is not None:
            groups.setdefault((record.location.get("page"), record.location.get("table"),
                               record.label, kind), []).append((record, value))
        if kind in hard_bounds and value is not None:
            low, high = hard_bounds[kind]
            if not low <= value <= high:
                findings.append(finding(
                    record, "physical_domain_violation", "high",
                    f"{kind} value {value:g} lies outside the generic physical domain [{low}, {high}]",
                    expected_domain=[low, high], proposed_value=None))
    for (_, _, _, kind), values in groups.items():
        if kind not in ("measurement", "count", "temperature") or len(values) < 8:
            continue
        nums = [value for _, value in values]
        median = statistics.median(nums)
        deviations = [abs(value - median) for value in nums]
        mad = statistics.median(deviations)
        if mad == 0:
            nonmodal = [(record, value) for record, value in values if value != median]
            # A constant column plus one departure is suspicious, but not an
            # ecological correction: report only extreme scale changes.
            threshold = max(10.0, abs(median) * 10)
            flagged = [(record, value) for record, value in nonmodal
                       if abs(value - median) > threshold]
        else:
            flagged = [(record, value) for record, value in values
                       if 0.6745 * abs(value - median) / mad > 6]
        for record, value in flagged:
            findings.append(finding(
                record, "within_form_numeric_outlier", "medium",
                f"{value:g} is a robust within-column outlier (median {median:g}, MAD {mad:g})",
                median=median, mad=mad, proposed_value=None))
    return findings


def taxonomy_findings(records, client, latitude=None, longitude=None):
    findings = []
    eligible = [record for record in records
                if (kind_of(record.label) == "species"
                    and record.location.get("kind") != "free_text"
                    and str(record.value or "").strip())]
    queries = list(dict.fromkeys(taxon_query(record.value) for record in eligible))
    cache = {}

    # Catalogue lookups are independent I/O. A small fixed pool preserves the
    # same calls and decisions while avoiding minutes of serial Fargate time on
    # long species checklists.
    workers = max(1, min(4, int(os.environ.get("FORMIDABLE_ECOLOGY_WORKERS", "4"))))
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(client.match, query): query for query in queries}
        for future in concurrent.futures.as_completed(futures):
            query = futures[future]
            try:
                cache[query.casefold()] = (*future.result(), None)
            except Exception as error:
                cache[query.casefold()] = (None, None, error)

    for record in eligible:
        original = " ".join(str(record.value).split())
        query = taxon_query(record.value)
        match, source, error = cache[query.casefold()]
        if error is not None:  # network failure is not a data finding
            findings.append(finding(record, "taxonomy_check_unavailable", "info",
                                    f"GBIF lookup failed: {type(error).__name__}", proposed_value=None))
            continue
        canonical_name = match.get("canonicalName")
        confidence = int(match.get("confidence") or 0)
        if not canonical_name or confidence < 90:
            findings.append(finding(
                record, "taxonomy_unmatched", "info",
                f"GBIF did not return a high-confidence taxon match (confidence {confidence})",
                gbif=match, source_url=source, proposed_value=None))
            continue
        comparison_query = query.strip(" \t.,;:")
        distance = edit_distance(comparison_query, canonical_name)
        # A large difference commonly means GBIF resolved a common name or
        # synonym to its scientific canonical name. That is useful matching
        # context, not evidence that the written value needs review.
        if comparison_query.casefold() != canonical_name.casefold() and distance <= 3:
            severity = "medium" if confidence >= 95 else "info"
            proposed = (canonical_name if query.casefold() == original.casefold()
                        else original.replace(query, canonical_name))
            findings.append(finding(
                record, "taxonomy_spelling_suggestion", severity,
                f"GBIF matched this name to {canonical_name!r} (confidence {confidence})",
                gbif=match, source_url=source,
                proposed_value=proposed if severity == "medium" else None,
                matched_input=query, edit_distance=distance))
        if latitude is not None and longitude is not None and match.get("usageKey"):
            try:
                occurrence, occurrence_url = client.nearby_count(
                    match["usageKey"], latitude, longitude)
            except Exception:
                continue
            if int(occurrence.get("count") or 0) == 0:
                findings.append(finding(
                    record, "no_nearby_gbif_records", "info",
                    "No GBIF occurrence records were found within approximately one degree; "
                    "absence of records is not evidence of species absence",
                    source_url=occurrence_url, proposed_value=None))
    return findings


def location_coordinates(records):
    """Find defensible coordinates already present on the form, if any.

    This is deliberately generic: separate latitude/longitude fields and a
    combined GPS/coordinate field are supported. No project/site lookup is
    used and ambiguous or physically invalid pairs are ignored.
    """
    latitude = longitude = None
    for record in records:
        kind = kind_of(record.label)
        value = number(record.value)
        if kind == "latitude" and value is not None and -90 <= value <= 90:
            latitude = value
        elif kind == "longitude" and value is not None and -180 <= value <= 180:
            longitude = value
    if latitude is not None and longitude is not None:
        return latitude, longitude
    for record in records:
        label = " ".join(str(record.label).casefold().split())
        if not any(term in label for term in ("gps", "coordinate", "location")):
            continue
        values = [float(value) for value in re.findall(
            r"(?<![A-Za-z0-9])[-+]?\d{1,3}(?:\.\d+)?(?![A-Za-z0-9])",
            str(record.value or "").replace(",", " "))]
        for first, second in zip(values, values[1:]):
            if -90 <= first <= 90 and -180 <= second <= 180:
                return first, second
    return None, None


def canonical_records(document):
    records = []
    for page in document.get("pages") or []:
        for field in page.get("metadata_fields") or []:
            records.append(Record({"page": page["page_number"], "field": field["id"],
                                   "bbox": field.get("bbox"),
                                   "xlsx_sheet": field.get("xlsx_sheet"),
                                   "xlsx_row": field.get("xlsx_row"),
                                   "xlsx_column": field.get("xlsx_column")},
                                  field["label"], field.get("value")))
        for item in page.get("free_text_regions") or []:
            records.append(Record({"page": page["page_number"], "field": item["id"],
                                   "kind": "free_text",
                                   "bbox": item.get("bbox"),
                                   "xlsx_sheet": item.get("xlsx_sheet"),
                                   "xlsx_row": item.get("xlsx_row"),
                                   "xlsx_column": item.get("xlsx_column")},
                                  item["label"], item.get("value")))
        for table in page.get("tables") or []:
            columns = table.get("columns") or []
            for row in table.get("rows") or []:
                for column, cell in zip(columns, row.get("cells") or []):
                    label = " ".join(x for x in (column.get("parent"), column.get("label")) if x)
                    records.append(Record(
                        {"page": page["page_number"], "table": table["id"],
                         "row": row["id"], "column": column["id"],
                         "bbox": cell.get("bbox"),
                         "xlsx_sheet": cell.get("xlsx_sheet"),
                         "xlsx_row": cell.get("xlsx_row"),
                         "xlsx_column": cell.get("xlsx_column")},
                        label, cell.get("value")))
    return records


def apply_findings(document, findings):
    lookup = {}
    for page in document.get("pages") or []:
        for field in page.get("metadata_fields") or []:
            lookup[(page["page_number"], None, None, None, field["id"])] = field
        for item in page.get("free_text_regions") or []:
            lookup[(page["page_number"], None, None, None, item["id"])] = item
        for table in page.get("tables") or []:
            for row in table.get("rows") or []:
                for cell in row.get("cells") or []:
                    lookup[(page["page_number"], table["id"], row["id"],
                            cell["column_id"], None)] = cell
    applied = 0
    for item in findings:
        loc = item["location"]
        key = (loc.get("page"), loc.get("table"), loc.get("row"),
               loc.get("column"), loc.get("field"))
        target = lookup.get(key)
        if target is not None:
            target.setdefault("ecology_flags", []).append(item)
            applied += 1
    return applied


def add_review_sheet(path, findings):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    if "ecology_review" in wb.sheetnames:
        del wb["ecology_review"]
    # Keep paper-page sheets first. The review UI indexes those sheets by page;
    # the audit sheet is useful for downloads but must never displace page 1.
    ws = wb.create_sheet("ecology_review")
    headers = ["severity", "code", "page", "table/field", "row", "column",
               "observed", "suggestion (not applied)", "message", "source"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for item in findings:
        loc = item["location"]
        ws.append([item["severity"], item["code"], loc.get("page"),
                   loc.get("table") or loc.get("field"), loc.get("row"),
                   loc.get("column"), item.get("observed"), item.get("proposed_value"),
                   item.get("message"), item.get("source_url")])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate((10, 30, 8, 24, 12, 22, 24, 28, 70, 55), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("canonical_json")
    parser.add_argument("--output", required=True)
    parser.add_argument("--cache", default=str(HERE / ".cache" / "gbif"))
    parser.add_argument("--latitude", type=float)
    parser.add_argument("--longitude", type=float)
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--annotated-canonical")
    parser.add_argument("--xlsx")
    args = parser.parse_args()
    document = json.loads(Path(args.canonical_json).read_text())
    records = canonical_records(document)
    findings = numeric_findings(records)
    if not args.offline:
        findings += taxonomy_findings(records, GBIFClient(Path(args.cache)),
                                      args.latitude, args.longitude)
    report = {"version": "formidable-ecology-review-v1", "source": args.canonical_json,
              "policy": "flags and suggestions only; extraction values are never changed",
              "records": len(records), "findings": findings}
    Path(args.output).write_text(json.dumps(report, indent=2, ensure_ascii=False))
    applied = 0
    if args.annotated_canonical or args.xlsx:
        applied = apply_findings(document, findings)
    if args.annotated_canonical:
        canonical.dump(document, args.annotated_canonical)
    if args.xlsx:
        canonical.write_xlsx(document, args.xlsx)
        add_review_sheet(args.xlsx, findings)
    print(json.dumps({"output": args.output, "records": len(records),
                      "findings": len(findings), "applied": applied}, indent=2))


if __name__ == "__main__":
    main()
