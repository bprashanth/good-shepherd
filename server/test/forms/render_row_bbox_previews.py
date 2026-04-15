"""Call /api/upload and render bbox overlays for matched Excel rows."""

from __future__ import annotations

import argparse
import mimetypes
from pathlib import Path
from io import BytesIO
from urllib.parse import urljoin
import base64

from PIL import Image, ImageDraw
import requests
from openpyxl import load_workbook

SYSTEM_SERIAL_HEADER = "(Good Shepherd) Row ID"
START_SERVER_HINT = (
    "good-shepherd/server$ uvicorn main:app --host 0.0.0.0 --port 8070 --reload"
)


def _normalize_server(server: str) -> str:
    """Ensure server string has a scheme and trailing slash."""
    server = server.strip()
    if not server.startswith("http://") and not server.startswith("https://"):
        server = f"http://{server}"
    return server.rstrip("/") + "/"


def _assert_server_available(base_url: str) -> None:
    """Verify API server is reachable before making upload request."""
    health_url = urljoin(base_url, "api/health")
    try:
        response = requests.get(health_url, timeout=3)
    except requests.RequestException:
        print(f"Server not reachable at {base_url}")
        print(f"Start it with: {START_SERVER_HINT}")
        raise SystemExit(1)
    if response.status_code != 200:
        print(f"Unexpected health response from {health_url}: {response.status_code}")
        print(f"Start it with: {START_SERVER_HINT}")
        raise SystemExit(1)


def _call_upload(base_url: str, image_path: Path) -> dict:
    """POST image to /api/upload and return JSON payload."""
    upload_url = urljoin(base_url, "api/upload")
    content_type = mimetypes.guess_type(str(image_path))[0] or "application/octet-stream"
    with image_path.open("rb") as fh:
        response = requests.post(
            upload_url,
            files={"image": (image_path.name, fh, content_type)},
            timeout=90,
        )
    if response.status_code != 200:
        raise RuntimeError(
            f"Upload failed ({response.status_code}) at {upload_url}: {response.text[:400]}"
        )
    return response.json()


def _extract_matching_serials(xlsx_bytes: bytes, match_word: str) -> set[int]:
    """Find all system serial values where any cell contains match_word."""
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    needle = match_word.casefold()

    serial_cols: set[int] = set()
    matched_serials: set[int] = set()

    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == SYSTEM_SERIAL_HEADER:
                serial_cols.add(cell.column)

        if not serial_cols:
            continue

        serial_value = None
        for col_idx in serial_cols:
            cell = next((c for c in row if c.column == col_idx), None)
            if cell is None or cell.value is None:
                continue
            try:
                serial_value = int(cell.value)
                break
            except (TypeError, ValueError):
                continue
        if serial_value is None:
            continue

        row_matches = any(
            cell.value is not None and needle in str(cell.value).casefold()
            for cell in row
            if cell.column not in serial_cols
        )
        if row_matches:
            matched_serials.add(serial_value)

    return matched_serials


def _draw_bbox(
    draw: ImageDraw.ImageDraw,
    image_size: tuple[int, int],
    bbox: dict,
    label: str,
    color: str,
) -> None:
    width, height = image_size
    x1 = int(bbox["left"] * width)
    y1 = int(bbox["top"] * height)
    x2 = int((bbox["left"] + bbox["width"]) * width)
    y2 = int((bbox["top"] + bbox["height"]) * height)
    draw.rectangle([x1, y1, x2, y2], outline=color, width=5)
    text_box = [x1, max(0, y1 - 24), x1 + 150, y1]
    draw.rectangle(text_box, fill=color)
    draw.text((x1 + 4, max(0, y1 - 20)), label, fill="white")


def render_previews(server: str, input_image: Path, match_word: str, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    base_url = _normalize_server(server)
    _assert_server_available(base_url)

    payload = _call_upload(base_url, input_image)
    rows = payload.get("rows", [])
    rows_by_serial = {
        int(r["system_serial"]): r.get("bbox")
        for r in rows
        if r.get("system_serial") is not None
    }
    xlsx = payload.get("xlsx")
    if not xlsx:
        raise RuntimeError("Upload payload missing xlsx")
    xlsx_bytes = base64.b64decode(xlsx)
    matched_serials = sorted(_extract_matching_serials(xlsx_bytes, match_word))

    image = Image.open(input_image).convert("RGB")
    draw = ImageDraw.Draw(image)

    drawn = 0
    for serial in matched_serials:
        bbox = rows_by_serial.get(serial)
        if not bbox:
            continue
        _draw_bbox(draw, image.size, bbox, f"id {serial}", "#D32F2F")
        drawn += 1

    footer = (
        f"match='{match_word}' | matched_serials={len(matched_serials)} | "
        f"drawn={drawn} | total_rows={len(rows)}"
    )
    draw.rectangle([12, 12, 980, 42], fill="#222222")
    draw.text((20, 18), footer, fill="white")

    output_name = f"{input_image.stem}_match_overlay.jpg"
    out_path = out_dir / output_name
    image.save(out_path, format="JPEG", quality=95)
    print(f"Wrote {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--server",
        type=str,
        default="0.0.0.0:8070",
        help="API server host:port or URL (default: 0.0.0.0:8070)",
    )
    parser.add_argument(
        "--input-image",
        type=Path,
        required=True,
        help="Path to one input form image",
    )
    parser.add_argument(
        "--match-word",
        type=str,
        required=True,
        help="Case-insensitive word/substring to match in returned Excel rows",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "bbox_previews",
        help="Directory where overlay JPEG is written",
    )
    args = parser.parse_args()
    if not args.input_image.exists():
        raise FileNotFoundError(f"Input image not found: {args.input_image}")
    render_previews(args.server, args.input_image, args.match_word, args.out_dir)


if __name__ == "__main__":
    main()
