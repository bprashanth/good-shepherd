"""
Converts extracted table data into an Excel file with confidence-based cell coloring.

Color rules:
  - confidence < 70%:  red border + light red fill
  - 70% <= conf < 85%: orange border + light orange fill
  - confidence >= 85%: no extra formatting

Universal fields appear as key-value rows above the data table.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side, Alignment


# --- Styles ---

_RED_SIDE = Side(style="thin", color="CC0000")
_ORANGE_SIDE = Side(style="thin", color="CC8800")

STYLE_RED = {
    "border": Border(left=_RED_SIDE, right=_RED_SIDE, top=_RED_SIDE, bottom=_RED_SIDE),
    "fill": PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
}
STYLE_ORANGE = {
    "border": Border(
        left=_ORANGE_SIDE, right=_ORANGE_SIDE,
        top=_ORANGE_SIDE, bottom=_ORANGE_SIDE,
    ),
    "fill": PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid"),
}
STYLE_HEADER = {
    "font": Font(bold=True),
    "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
STYLE_UF_KEY = {
    "font": Font(bold=True, color="333333"),
}
STYLE_LEGEND = {
    "font": Font(italic=True, color="666666", size=9),
}


def _apply_confidence_style(cell, confidence: float):
    """Apply red/orange styling based on confidence score (0-100 scale)."""
    if confidence < 70:
        cell.border = STYLE_RED["border"]
        cell.fill = STYLE_RED["fill"]
    elif confidence < 85:
        cell.border = STYLE_ORANGE["border"]
        cell.fill = STYLE_ORANGE["fill"]


def to_xlsx(extracted: dict) -> bytes:
    """
    Convert extracted table data to an Excel file.

    Args:
        extracted: dict from table_extractor.extract(), with keys:
            - tables: list of table dicts (column_headers, data_rows)
            - key_value_pairs: list of {key, value, key_confidence, value_confidence}

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Data"
    current_row = 1

    # --- Legend ---
    ws.cell(row=current_row, column=1, value="Confidence legend:")
    ws.cell(row=current_row, column=1).font = STYLE_LEGEND["font"]

    legend_red = ws.cell(row=current_row, column=2, value="Red = low (<70%)")
    legend_red.font = Font(italic=True, color="CC0000", size=9)
    _apply_confidence_style(legend_red, 50)

    legend_orange = ws.cell(row=current_row, column=3, value="Orange = mid (70-85%)")
    legend_orange.font = Font(italic=True, color="CC8800", size=9)
    _apply_confidence_style(legend_orange, 75)

    legend_ok = ws.cell(row=current_row, column=4, value="No highlight = high (85%+)")
    legend_ok.font = STYLE_LEGEND["font"]

    current_row += 2

    # --- Universal fields (key-value pairs) ---
    kv_pairs = extracted.get("key_value_pairs", [])
    if kv_pairs:
        for kv in kv_pairs:
            key_cell = ws.cell(row=current_row, column=1, value=kv["key"])
            key_cell.font = STYLE_UF_KEY["font"]
            _apply_confidence_style(key_cell, kv["key_confidence"])

            val_cell = ws.cell(row=current_row, column=2, value=kv["value"])
            _apply_confidence_style(val_cell, kv["value_confidence"])

            current_row += 1
        current_row += 1  # blank row after universal fields

    # --- Tables ---
    for table in extracted.get("tables", []):
        headers = table.get("column_headers", [])
        data_rows = table.get("data_rows", [])

        # Build ordered header list by column_index
        header_list = sorted(headers, key=lambda h: h["column_index"])
        header_texts = [h["text"] for h in header_list]

        # Write header row
        for col_idx, h in enumerate(header_list, 1):
            hcell = ws.cell(row=current_row, column=col_idx, value=h["text"])
            hcell.font = STYLE_HEADER["font"]
            hcell.fill = STYLE_HEADER["fill"]
        current_row += 1

        # Write data rows
        for row_data in data_rows:
            cells_by_header = {}
            for c in row_data.get("_cells", []):
                cells_by_header[c["header"]] = c

            for col_idx, header_text in enumerate(header_texts, 1):
                cell_info = cells_by_header.get(header_text)
                if cell_info:
                    dcell = ws.cell(
                        row=current_row, column=col_idx, value=cell_info["text"],
                    )
                    _apply_confidence_style(dcell, cell_info["confidence"])
                else:
                    ws.cell(row=current_row, column=col_idx, value="")
            current_row += 1

        current_row += 1  # blank row between tables

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
